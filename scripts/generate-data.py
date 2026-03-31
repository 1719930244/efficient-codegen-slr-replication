#!/usr/bin/env python3
"""
Generate replication package data from primary study sources.

Uses fine-grained-classification.csv as the authoritative study list (125 entries),
excludes only DUPLICATE entries (3) to get 122 primary studies.
Enriches with metadata from primary-studies-assessment.xlsx where available.
"""

import csv
import json
import os
import sys

try:
    import openpyxl
except ImportError:
    print("pip install openpyxl")
    sys.exit(1)

SURVEY_DIR = os.path.expanduser("~/overleaf/efficient-codegen-survey")
DATA_DIR = os.path.join(SURVEY_DIR, "data")
OUTPUT_DIR = os.path.join(os.path.dirname(__file__), "..", "data")

CATEGORY_NAMES = {
    "1a": "Data Selection",
    "1b": "Data Quality Assessment & Synthesis",
    "1c": "Data Deduplication",
    "1d": "Data Mixing",
    "2a": "Efficient Pre-training",
    "2b": "Parameter-Efficient Fine-Tuning (PEFT)",
    "2c": "Knowledge Distillation",
    "2d": "Curriculum Learning",
    "2e": "Training-time Compression",
    "2f": "RL-based Training",
    "3a": "Speculative Decoding",
    "3b": "Early Exit",
    "3c": "Non-Autoregressive Generation",
    "3d": "Prompt Compression",
    "3e": "Context Pruning",
    "3f": "KV Cache Optimization",
    "3g": "Post-Training Quantization",
    "3h": "Model Pruning",
    "3i": "Model Routing",
    "3j": "Adaptive Sampling",
    "3k": "Multi-Agent Orchestration",
    "3l": "Code-Specific Optimization",
    "3m": "System-Level Serving",
    "3n": "Prompt Engineering",
    "3o": "Chain-of-Thought Optimization",
    "4a": "Benchmark Design",
    "4b": "Empirical Study",
}

RQ_NAMES = {
    "RQ1": "Data Preparation",
    "RQ2": "Model Training",
    "RQ3": "Inference Optimization",
    "RQ4": "Evaluation",
}


def load_csv_studies():
    """Load primary studies from classification CSV, excluding only DUPLICATE entries."""
    with open(os.path.join(DATA_DIR, "fine-grained-classification.csv")) as f:
        rows = list(csv.DictReader(f))

    included = [
        r for r in rows
        if "DUPLICATE" not in r.get("scope_flags", "")
    ]
    return {r["key"]: r for r in included}


def load_xlsx_metadata():
    """Load metadata from assessment xlsx for enrichment."""
    xlsx_path = os.path.join(DATA_DIR, "primary-studies-assessment.xlsx")
    if not os.path.exists(xlsx_path):
        return {}
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb["Primary Studies"]
    meta = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[1]:
            continue
        key = row[1]
        meta[key] = {
            "author": row[3] or "",
            "venue_type": row[6] or "",
            "source": row[7] or "",
            "qa1": row[14],
            "qa2": row[15],
            "qa3": row[16],
            "qa4": row[17],
            "qa_total": row[18],
        }
    return meta


def build_studies():
    """Build the complete primary study set from CSV + xlsx metadata."""
    csv_data = load_csv_studies()
    xlsx_meta = load_xlsx_metadata()

    studies = {}
    for key, row in csv_data.items():
        meta = xlsx_meta.get(key, {})
        studies[key] = {
            "key": key,
            "title": row["title"],
            "author": meta.get("author", ""),
            "year": int(row["year"]),
            "venue": row["venue"],
            "venue_type": meta.get("venue_type", ""),
            "source": meta.get("source", ""),
            "rq": row["new_rq"],
            "primary_categories": row["primary_categories"],
            "secondary_categories": row.get("secondary_categories", ""),
            "scope_flags": row.get("scope_flags", ""),
            "brief_rationale": row.get("brief_rationale", ""),
            "qa_total": meta.get("qa_total", ""),
        }
    return studies


def parse_rqs(rq_str):
    if not rq_str:
        return []
    return [r.strip() for r in rq_str.split(";")]


def parse_categories(cat_str):
    if not cat_str or cat_str == "None":
        return []
    return [c.strip() for c in cat_str.split(";")]


def get_category_code(cat):
    return cat.split("-")[0] if "-" in cat else cat


def write_primary_studies_csv(studies):
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    outpath = os.path.join(OUTPUT_DIR, "primary-studies.csv")

    sorted_studies = sorted(studies.values(), key=lambda s: (s["year"], s["key"]))

    with open(outpath, "w", newline="") as f:
        writer = csv.writer(f)
        writer.writerow([
            "ID", "Key", "Title", "Year", "Venue", "Venue Type", "Source",
            "RQ", "Primary Categories", "Secondary Categories",
            "Scope Flags", "Brief Rationale"
        ])
        for i, s in enumerate(sorted_studies, 1):
            writer.writerow([
                f"S{i:03d}",
                s["key"],
                s["title"],
                s["year"],
                s["venue"],
                s["venue_type"],
                s["source"],
                s["rq"],
                s["primary_categories"],
                s["secondary_categories"],
                s["scope_flags"],
                s["brief_rationale"],
            ])

    print(f"Written {len(sorted_studies)} studies to {outpath}")
    return sorted_studies


def write_by_rq(studies):
    rq_dir = os.path.join(OUTPUT_DIR, "by-rq")
    os.makedirs(rq_dir, exist_ok=True)

    rq_studies = {"RQ1": [], "RQ2": [], "RQ3": [], "RQ4": []}

    for s in studies:
        for rq in parse_rqs(s["rq"]):
            if rq in rq_studies:
                rq_studies[rq].append(s)

    for rq, items in rq_studies.items():
        outpath = os.path.join(rq_dir, f"{rq.lower()}-studies.csv")
        with open(outpath, "w", newline="") as f:
            writer = csv.writer(f)
            writer.writerow(["Key", "Title", "Year", "Venue", "Categories"])
            for s in sorted(items, key=lambda x: (x["year"], x["key"])):
                writer.writerow([
                    s["key"], s["title"], s["year"], s["venue"],
                    s["primary_categories"],
                ])
        print(f"  {rq}: {len(items)} studies -> {outpath}")


def write_classification_scheme(studies):
    outpath = os.path.join(OUTPUT_DIR, "classification-scheme.csv")

    cat_studies = {}
    for s in studies:
        for cat in parse_categories(s["primary_categories"]):
            code = get_category_code(cat)
            if code not in cat_studies:
                cat_studies[code] = []
            cat_studies[code].append(s["key"])

        for cat in parse_categories(s.get("secondary_categories", "")):
            code = get_category_code(cat)
            if code not in cat_studies:
                cat_studies[code] = []
            if s["key"] not in cat_studies[code]:
                cat_studies[code].append(s["key"])

    with open(outpath, "w", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(["Code", "Category", "RQ", "Study Count", "Study Keys"])
        for code in sorted(cat_studies.keys()):
            rq = f"RQ{code[0]}"
            name = CATEGORY_NAMES.get(code, code)
            keys = sorted(cat_studies[code])
            writer.writerow([code, name, rq, len(keys), "; ".join(keys)])

    print(f"Written classification scheme to {outpath}")


def write_statistics(studies):
    from collections import Counter

    stats = {
        "total_studies": len(studies),
        "year_distribution": dict(sorted(Counter(s["year"] for s in studies).items())),
        "rq_distribution": {},
        "cross_rq_studies": sum(1 for s in studies if len(parse_rqs(s["rq"])) > 1),
        "category_distribution": {},
    }

    rq_counts = Counter()
    for s in studies:
        for rq in parse_rqs(s["rq"]):
            rq_counts[rq] += 1
    stats["rq_distribution"] = dict(sorted(rq_counts.items()))

    cat_counts = Counter()
    for s in studies:
        for cat in parse_categories(s["primary_categories"]):
            code = get_category_code(cat)
            name = CATEGORY_NAMES.get(code, code)
            cat_counts[f"{code}-{name}"] += 1
    stats["category_distribution"] = dict(sorted(cat_counts.items()))

    outpath = os.path.join(OUTPUT_DIR, "statistics.json")
    with open(outpath, "w") as f:
        json.dump(stats, f, indent=2, sort_keys=False)
    print(f"Written statistics to {outpath}")


def write_reporting_compliance(studies):
    """Write reporting compliance data from xlsx QA scores."""
    xlsx_path = os.path.join(DATA_DIR, "primary-studies-assessment.xlsx")
    if not os.path.exists(xlsx_path):
        print("Skipping reporting compliance (no xlsx)")
        return

    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb["Primary Studies"]
    headers = [cell.value for cell in ws[1]]

    study_keys = {s["key"] for s in studies}
    compliance = {}

    for row in ws.iter_rows(min_row=2, values_only=True):
        key = row[1]
        if not key or key not in study_keys:
            continue
        compliance[key] = {
            "qa1_empirical_evaluation": row[14],
            "qa2_baseline_comparison": row[15],
            "qa3_experimental_config": row[16],
            "qa4_reproducibility_artifacts": row[17],
            "qa_total": row[18],
        }

    outpath = os.path.join(OUTPUT_DIR, "reporting-compliance.json")
    with open(outpath, "w") as f:
        json.dump(compliance, f, indent=2, sort_keys=True)
    print(f"Written reporting compliance for {len(compliance)} studies to {outpath}")


def main():
    all_studies = build_studies()
    print(f"Total primary studies: {len(all_studies)}")

    sorted_studies = write_primary_studies_csv(all_studies)
    write_by_rq(sorted_studies)
    write_classification_scheme(sorted_studies)
    write_statistics(sorted_studies)
    write_reporting_compliance(sorted_studies)


if __name__ == "__main__":
    main()
